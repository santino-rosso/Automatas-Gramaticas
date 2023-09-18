import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os

# Exporta los resultados a un archivo Excel
def exportar_a_excel(trafico_por_usuario, trafico_maximo):
    workbook = openpyxl.Workbook() #Creamos el archivo excel
    sheet = workbook.active
    sheet.title = "Resultados"

    # Encabezados
    encabezados = ["Usuario", "Tráfico (bytes)"]
    for col, encabezado in enumerate(encabezados, start=1):
        celda = sheet.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True) #Texto en negrita
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for row, (usuario, trafico) in enumerate(trafico_por_usuario.items(), start=2):
        sheet.cell(row=row, column=1, value=usuario)
        sheet.cell(row=row, column=2, value=trafico)

    # Resaltar el usuario con más tráfico
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == trafico_maximo:
                cell.fill = PatternFill(start_color='6DC36D', end_color='6DC36D', fill_type='solid')

    # Guardar archivo
    archivo_salida = "resultados_trafico.xlsx"

    # Luego de terminar, verifica si el archivo existe y elimínalo si es necesario
    if os.path.exists(archivo_salida):
        os.remove(archivo_salida)
    
    workbook.save(archivo_salida)
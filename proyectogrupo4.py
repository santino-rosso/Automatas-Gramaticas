import re
from datetime import datetime
import time
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Define las expresiones regulares para validar campos

def Validar_ID(string):
    reg = re.compile(r"^\d{6,7}$")
    return bool(reg.match(string))

def Validar_ID_Sesion(string):
    reg = re.compile(r"^(([0-9]|[A-F]){8}-?([0-9]|[A-F]){8})$")
    return bool(reg.match(string))

def Validar_ID_Conexion_Unico(string):
    reg = re.compile(r"^([0-9]|[a-f]){16}$")
    return bool(reg.match(string))

def Validar_Usuario(string):
    reg = re.compile(r"^.*\D+.*$")
    return bool(reg.match(string))

def Validar_Inicio_Conexion_Dia(string):
    reg = re.compile(r"^(2019|202[0-3])-(0[1-9]|1[0-2])-(0[1-9]|[1-2][0-9]|3[0-1])$")
    return bool(reg.match(string))

def Validar_Inicio_Conexion_Hora(string):
    reg = re.compile(r"^([0-1][0-9]|[2][0-3]):([0-5][0-9]):([0-5][0-9])$")
    return bool(reg.match(string))

def Validar_Fin_Conexion_Dia(string):
    reg = re.compile(r"^(2019|202[0-3])-(0[1-9]|1[0-2])-(0[1-9]|[1-2][0-9]|3[0-1])$")
    return bool(reg.match(string))

def Validar_Fin_Conexion_Hora(string):
    reg = re.compile(r"^([0-1][0-9]|[2][0-3]):([0-5][0-9]):([0-5][0-9])$")
    return bool(reg.match(string))

def Validar_Session_Time(string):
    reg = re.compile(r"^\d+$")
    return bool(reg.match(string))

def Validar_Input_Octects(string):
    reg = re.compile(r"^\d+$")
    return bool(reg.match(string))

def Validar_Output_Octects(string):
    reg = re.compile(r"^\d+$")
    return bool(reg.match(string))

def Validar_MAC_AP(string):
    reg = re.compile(r"^(([A-F]|[0-9]){2}-){5}([A-F]|[0-9]){2}:[A-Z]{4}$")
    return bool(reg.match(string))

def Validar_MAC_Cliente(string):
    reg = re.compile(r"^(([A-F]|[0-9]){2}-){5}([A-F]|[0-9]){2}$")
    return bool(reg.match(string))

# Lee el archivo de entrada y filtra las líneas válidas
def leer_archivo(nombre_archivo):
    lineas_validas = []
    with open(nombre_archivo, 'r') as archivo:
        lineas = archivo.readlines()
        for linea in lineas[1:]:  # Ignorar la primera línea de encabezado
            campos = linea.strip().split(",")
            campos.pop()
            campos.pop()
            if (
                len(campos) == 16 and
                Validar_ID(campos[0]) and
                Validar_ID_Sesion(campos[1]) and
                Validar_ID_Conexion_Unico(campos[2]) and
                Validar_Usuario(campos[3]) and 
                Validar_Inicio_Conexion_Dia(campos[6]) and 
                Validar_Inicio_Conexion_Hora(campos[7]) and
                Validar_Fin_Conexion_Dia(campos[8]) and 
                Validar_Fin_Conexion_Hora(campos[9]) and
                Validar_Session_Time(campos[10]) and
                Validar_Input_Octects(campos[11]) and
                Validar_Output_Octects(campos[12]) and
                Validar_MAC_AP(campos[13]) and
                Validar_MAC_Cliente(campos[14])
            ):
                lineas_validas.append(campos)
    return lineas_validas

def calcular_trafico(lineas, fecha_inicio, fecha_fin):
    trafico_por_usuario = {}
    # Define las columnas correspondientes a cada campo
    columna_fecha_inicio = 6
    columna_fecha_hora_inicio = 7
    columna_fecha_fin = 8
    columna_fecha_hora_fin = 9
    columna_usuario = 3
    columna_input_octets = 11
    columna_output_octets = 12
    
    for linea in lineas:
        # Obtén los valores de las columnas para esta línea
        fecha_inicio_str = linea[columna_fecha_inicio]
        fecha_hora_inicio_str = linea[columna_fecha_hora_inicio]
        fecha_fin_str = linea[columna_fecha_fin]
        fecha_hora_fin_str = linea[columna_fecha_hora_fin]
        usuario = linea[columna_usuario]
        input_octets = int(linea[columna_input_octets])
        output_octets = int(linea[columna_output_octets])
        
        # Convierte las fechas y horas en objetos datetime
        fecha_inicio_conexion = datetime.strptime(fecha_inicio_str + ' ' + fecha_hora_inicio_str, '%Y-%m-%d %H:%M:%S')
        fecha_fin_conexion = datetime.strptime(fecha_fin_str + ' ' + fecha_hora_fin_str, '%Y-%m-%d %H:%M:%S')
        
        # Verifica si la conexión está en el rango de fechas especificado
        if fecha_inicio <= fecha_inicio_conexion <= fecha_fin or fecha_inicio <= fecha_fin_conexion <= fecha_fin:
            trafico_total = input_octets + output_octets
            
            # Agrega el tráfico al usuario correspondiente en el diccionario
            if usuario in trafico_por_usuario:
                trafico_por_usuario[usuario] += trafico_total
            else:
                trafico_por_usuario[usuario] = trafico_total
                
    return trafico_por_usuario

# Encuentra el usuario con más tráfico
def encontrar_usuario_max_trafico(trafico_por_usuario):
    lista_valores = trafico_por_usuario.values()
    trafico_maximo = max(lista_valores)
    usuario_max_trafico = None
    for usuario, trafico in trafico_por_usuario.items():
        if trafico == trafico_maximo:
            usuario_max_trafico = usuario
    return usuario_max_trafico, trafico_maximo


# Exporta los resultados a un archivo Excel
def exportar_a_excel(trafico_por_usuario, usuario_max_trafico, trafico_maximo):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"

    # Encabezados
    encabezados = ["Usuario", "Tráfico (bytes)"]
    for col, encabezado in enumerate(encabezados, 1):
        celda = sheet.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for row, (usuario, trafico) in enumerate(trafico_por_usuario.items(), start=2):
        sheet.cell(row=row, column=1, value=usuario)
        sheet.cell(row=row, column=2, value=trafico)

    # Resaltar el usuario con más tráfico
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == trafico_maximo:
                cell.fill = PatternFill(start_color='6DC36D', end_color='6DC36D', fill_type='solid')

    # Guardar archivo
    archivo_salida = "resultados_trafico.xlsx"

    # Luego de terminar, verifica si el archivo existe y elimínalo si es necesario
    if os.path.exists(archivo_salida):
        os.remove(archivo_salida)
    
    time.sleep(3)
    workbook.save(archivo_salida)


def main():
    while True:
        # Lee el archivo y filtra las líneas válidas
        archivo_entrada = "basededatos.csv"  # Reemplaza con la ruta de tu archivo de entrada
        lineas_validas = leer_archivo(archivo_entrada)

        if not lineas_validas:
            print("No se encontraron datos válidos en el archivo de entrada.")
            return

        # Solicita al usuario las fechas de inicio y fin
        fecha_inicio = input("Ingrese la fecha de inicio (yyyy-mm-dd HH:mm:ss): ")
        fecha_fin = input("Ingrese la fecha de fin (yyyy-mm-dd HH:mm:ss): ")

        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d %H:%M:%S")
            fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            print("Formato de fecha incorrecto. Utilice yyyy-mm-dd HH:mm:ss.")
            return

        # Calcula el tráfico total para cada usuario en el rango de fechas
        trafico_por_usuario = calcular_trafico(lineas_validas, fecha_inicio, fecha_fin)

        if not trafico_por_usuario:
            print("No se encontraron datos en el rango de fechas especificado.")
            return

        # Encuentra el usuario con más tráfico
        usuario_max_trafico, trafico_maximo = encontrar_usuario_max_trafico(trafico_por_usuario)
        
        print("Calculando")
        time.sleep(1.5)

        print(f"Usuario con más tráfico en el rango de fechas: {usuario_max_trafico}")
        print(f"Tráfico máximo: {trafico_maximo} bytes")
        print("Deseas exportar los datos a un excel? (y/n)")
        exportar = input()
        if exportar.lower() == "y":
            exportar_a_excel(trafico_por_usuario, usuario_max_trafico, trafico_maximo)
            print(f"Resultados exportados a 'resultados_trafico.xlsx'.")
        
        print("Deseas volver a realizar trafico de datos? (y/n)")
        volver_a_iniciar = input()
        if volver_a_iniciar.lower() == "n":
            time.sleep(1.5)
            print("Finalizado!")
            break
        
if __name__ == "__main__":
    main()

